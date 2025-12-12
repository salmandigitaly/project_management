from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime


def generate_excel_template() -> Workbook:
    """Generate Excel template for bulk import"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Projects
    ws_projects = wb.create_sheet("Projects")
    headers = ["key", "name", "description", "platform", "start_date", "end_date", "project_lead_email"]
    ws_projects.append(headers)
    
    # Style headers
    for cell in ws_projects[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add example row
    ws_projects.append([
        "PROJ",
        "HRMS Project",
        "Human Resource Management System",
        "web",
        "2025-01-01",
        "2025-12-31",
        "lead@company.com"
    ])
    
    # Sheet 2: Epics
    ws_epics = wb.create_sheet("Epics")
    headers = ["project_key", "name", "description", "start_date", "end_date"]
    ws_epics.append(headers)
    
    for cell in ws_epics[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    ws_epics.append([
        "PROJ",
        "User Management",
        "User authentication and authorization",
        "2025-01-01",
        "2025-03-31"
    ])
    
    # Sheet 3: Sprints
    ws_sprints = wb.create_sheet("Sprints")
    headers = ["project_key", "name", "goal", "start_date", "end_date"]
    ws_sprints.append(headers)
    
    for cell in ws_sprints[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    ws_sprints.append([
        "PROJ",
        "Sprint 1",
        "Complete login and registration",
        "2025-01-01",
        "2025-01-15"
    ])
    
    # Sheet 4: Features
    ws_features = wb.create_sheet("Features")
    headers = ["project_key", "epic_name", "name", "description", "priority", "status"]
    ws_features.append(headers)
    
    for cell in ws_features[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")
    
    ws_features.append([
        "PROJ",
        "User Management",
        "Login Module",
        "Complete login functionality",
        "high",
        "todo"
    ])
    
    # Sheet 5: Issues
    ws_issues = wb.create_sheet("Issues")
    headers = [
        "project_key", "epic_name", "feature_name", "sprint_name", "type", "name",
        "description", "priority", "status", "assignee_email",
        "story_points", "estimated_hours"
    ]
    ws_issues.append(headers)
    
    for cell in ws_issues[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws_issues.append([
        "PROJ",
        "User Management",
        "Login Module",
        "Sprint 1",
        "story",
        "Login Feature",
        "Implement user login with email and password",
        "high",
        "todo",
        "dev@company.com",
        5,
        8.0
    ])
    
    # Sheet 6: Subtasks
    ws_subtasks = wb.create_sheet("Subtasks")
    headers = [
        "project_key", "parent_issue_name", "name", "description",
        "priority", "status", "assignee_email", "estimated_hours"
    ]
    ws_subtasks.append(headers)
    
    for cell in ws_subtasks[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    ws_subtasks.append([
        "PROJ",
        "Login Feature",
        "Design login UI",
        "Create mockups and wireframes",
        "medium",
        "todo",
        "designer@company.com",
        2.0
    ])
    
    return wb
