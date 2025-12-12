"""
Script to generate a sample Excel file with realistic project data for bulk import testing
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

def create_sample_excel():
    """Create a sample Excel file with realistic project management data"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # ==================== PROJECTS SHEET ====================
    ws_projects = wb.create_sheet("Projects")
    project_headers = ["key", "name", "description", "platform", "start_date", "end_date", "project_lead_email"]
    ws_projects.append(project_headers)
    
    # Style headers
    for cell in ws_projects[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add sample projects
    projects_data = [
        ["HRMS", "HR Management System", "Complete HR management solution with employee tracking, payroll, and attendance", "web", "2025-01-01", "2025-12-31", "admin@company.com"],
        ["ECOM", "E-Commerce Platform", "Online shopping platform with payment integration and inventory management", "web", "2025-02-01", "2025-11-30", "admin@company.com"],
    ]
    
    for project in projects_data:
        ws_projects.append(project)
    
    # ==================== EPICS SHEET ====================
    ws_epics = wb.create_sheet("Epics")
    epic_headers = ["project_key", "name", "description", "start_date", "end_date"]
    ws_epics.append(epic_headers)
    
    for cell in ws_epics[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Add sample epics
    epics_data = [
        ["HRMS", "User Management", "User authentication, authorization, and profile management", "2025-01-01", "2025-03-31"],
        ["HRMS", "Employee Management", "Employee records, onboarding, and offboarding", "2025-02-01", "2025-05-31"],
        ["HRMS", "Attendance Tracking", "Clock in/out, leave management, and attendance reports", "2025-04-01", "2025-07-31"],
        ["ECOM", "Product Catalog", "Product listing, categories, and search functionality", "2025-02-01", "2025-04-30"],
        ["ECOM", "Shopping Cart", "Cart management, checkout, and order processing", "2025-03-01", "2025-06-30"],
    ]
    
    for epic in epics_data:
        ws_epics.append(epic)
    
    # ==================== SPRINTS SHEET ====================
    ws_sprints = wb.create_sheet("Sprints")
    sprint_headers = ["project_key", "name", "goal", "start_date", "end_date"]
    ws_sprints.append(sprint_headers)
    
    for cell in ws_sprints[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Add sample sprints
    sprints_data = [
        ["HRMS", "Sprint 1", "Complete user authentication and basic profile", "2025-01-01", "2025-01-14"],
        ["HRMS", "Sprint 2", "Employee CRUD operations and search", "2025-01-15", "2025-01-28"],
        ["HRMS", "Sprint 3", "Attendance clock in/out functionality", "2025-01-29", "2025-02-11"],
        ["ECOM", "Sprint 1", "Product listing and category management", "2025-02-01", "2025-02-14"],
        ["ECOM", "Sprint 2", "Shopping cart and checkout flow", "2025-02-15", "2025-02-28"],
    ]
    
    for sprint in sprints_data:
        ws_sprints.append(sprint)
    
    # ==================== FEATURES SHEET ====================
    ws_features = wb.create_sheet("Features")
    feature_headers = ["project_key", "epic_name", "name", "description", "priority", "status"]
    ws_features.append(feature_headers)
    
    for cell in ws_features[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")
    
    # Add sample features
    features_data = [
        ["HRMS", "User Management", "Login Module", "Email/password login with JWT authentication", "highest", "todo"],
        ["HRMS", "User Management", "Registration Module", "User registration with email verification", "high", "todo"],
        ["HRMS", "User Management", "Profile Management", "User profile viewing and editing", "medium", "todo"],
        ["HRMS", "Employee Management", "Employee CRUD", "Create, read, update, delete employee records", "highest", "todo"],
        ["HRMS", "Employee Management", "Employee Search", "Search and filter employees by various criteria", "high", "todo"],
        ["HRMS", "Attendance Tracking", "Clock In/Out", "Employee time tracking with clock in/out", "highest", "todo"],
        ["ECOM", "Product Catalog", "Product Listing", "Display products with images and details", "highest", "todo"],
        ["ECOM", "Shopping Cart", "Cart Management", "Add, remove, update items in cart", "highest", "todo"],
    ]
    
    for feature in features_data:
        ws_features.append(feature)
    
    # ==================== ISSUES SHEET ====================
    ws_issues = wb.create_sheet("Issues")
    issue_headers = [
        "project_key", "epic_name", "feature_name", "sprint_name", "type", "name",
        "description", "priority", "status", "assignee_email", "story_points", "estimated_hours"
    ]
    ws_issues.append(issue_headers)
    
    for cell in ws_issues[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Add sample issues
    issues_data = [
        ["HRMS", "User Management", "Login Module", "Sprint 1", "story", "Implement Login API", "Create POST /auth/login endpoint with JWT token generation", "highest", "todo", "dev@company.com", 5, 8.0],
        ["HRMS", "User Management", "Login Module", "Sprint 1", "story", "Create Login UI", "Design and implement login form with validation", "high", "todo", "frontend@company.com", 3, 6.0],
        ["HRMS", "User Management", "Registration Module", "Sprint 1", "story", "User Registration API", "Create POST /auth/register endpoint with email verification", "high", "todo", "dev@company.com", 5, 10.0],
        ["HRMS", "Employee Management", "Employee CRUD", "Sprint 2", "story", "Create Employee API", "POST endpoint to create new employee records", "highest", "todo", "dev@company.com", 3, 5.0],
        ["HRMS", "Employee Management", "Employee CRUD", "Sprint 2", "story", "List Employees API", "GET endpoint with pagination and filtering", "high", "todo", "dev@company.com", 3, 4.0],
        ["HRMS", "Employee Management", "Employee Search", "Sprint 2", "task", "Implement Search Algorithm", "Add search functionality for employee names and IDs", "medium", "todo", "dev@company.com", 2, 3.0],
        ["HRMS", "Attendance Tracking", "Clock In/Out", "Sprint 3", "story", "Clock In API", "POST endpoint to record clock in time", "highest", "todo", "dev@company.com", 3, 5.0],
        ["HRMS", "Attendance Tracking", "Clock In/Out", "Sprint 3", "story", "Clock Out API", "POST endpoint to record clock out time and calculate hours", "highest", "todo", "dev@company.com", 3, 5.0],
        ["ECOM", "Product Catalog", "Product Listing", "Sprint 1", "story", "Product API", "CRUD endpoints for product management", "highest", "todo", "dev@company.com", 5, 8.0],
        ["ECOM", "Product Catalog", "Product Listing", "Sprint 1", "story", "Product Display UI", "Grid view of products with images and prices", "high", "todo", "frontend@company.com", 5, 10.0],
        ["ECOM", "Shopping Cart", "Cart Management", "Sprint 2", "story", "Add to Cart API", "POST endpoint to add items to cart", "highest", "todo", "dev@company.com", 3, 5.0],
        ["ECOM", "Shopping Cart", "Cart Management", "Sprint 2", "story", "Cart UI", "Display cart items with quantity controls", "high", "todo", "frontend@company.com", 3, 6.0],
    ]
    
    for issue in issues_data:
        ws_issues.append(issue)
    
    # ==================== SUBTASKS SHEET ====================
    ws_subtasks = wb.create_sheet("Subtasks")
    subtask_headers = [
        "project_key", "parent_issue_name", "name", "description",
        "priority", "status", "assignee_email", "estimated_hours"
    ]
    ws_subtasks.append(subtask_headers)
    
    for cell in ws_subtasks[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    # Add sample subtasks
    subtasks_data = [
        ["HRMS", "Implement Login API", "Design database schema", "Create users table with necessary fields", "high", "todo", "dev@company.com", 2.0],
        ["HRMS", "Implement Login API", "Write authentication logic", "Implement password hashing and JWT generation", "highest", "todo", "dev@company.com", 4.0],
        ["HRMS", "Implement Login API", "Add input validation", "Validate email format and password strength", "medium", "todo", "dev@company.com", 1.0],
        ["HRMS", "Create Login UI", "Design mockup", "Create UI mockup in Figma", "high", "todo", "designer@company.com", 2.0],
        ["HRMS", "Create Login UI", "Implement form", "Build login form with React", "high", "todo", "frontend@company.com", 3.0],
        ["HRMS", "Create Employee API", "Design employee schema", "Define employee model with all required fields", "high", "todo", "dev@company.com", 1.5],
        ["HRMS", "Create Employee API", "Implement validation", "Add validation for employee data", "medium", "todo", "dev@company.com", 1.5],
        ["ECOM", "Product API", "Design product schema", "Create product model with categories", "high", "todo", "dev@company.com", 2.0],
        ["ECOM", "Product API", "Add image upload", "Implement image upload functionality", "medium", "todo", "dev@company.com", 3.0],
    ]
    
    for subtask in subtasks_data:
        ws_subtasks.append(subtask)
    
    # Save the workbook
    filename = "bulk_import_sample.xlsx"
    wb.save(filename)
    print(f"✅ Sample Excel file created: {filename}")
    print(f"\nFile contains:")
    print(f"  - {len(projects_data)} Projects")
    print(f"  - {len(epics_data)} Epics")
    print(f"  - {len(sprints_data)} Sprints")
    print(f"  - {len(features_data)} Features")
    print(f"  - {len(issues_data)} Issues")
    print(f"  - {len(subtasks_data)} Subtasks")
    print(f"\nYou can now upload this file to test the bulk import feature!")

if __name__ == "__main__":
    create_sample_excel()
